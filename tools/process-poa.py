import csv
import json
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
POA_PATH = ROOT / "capas_oficiales_2025" / "metadata" / "POA 01-10-2024-2025.xlsx"
METADATA_DIR = ROOT / "capas_oficiales_2025" / "metadata"
DATA_DIR = ROOT / "data"


def clean_value(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def slug(value):
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")


def match_key(value):
    text = slug(value)
    text = re.sub(r"^(torre|puesto|base|brigada)_+", "", text)
    return text


def write_csv(path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        return
    headers = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def row_to_dict(headers, values):
    output = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        output[header] = clean_value(values[index] if index < len(values) else "")
    return output


def export_terrestre(ws):
    headers = [clean_value(cell.value) for cell in ws[4]]
    headers = headers[: max(index for index, header in enumerate(headers) if header) + 1]
    rows = []
    for excel_row in ws.iter_rows(min_row=5, values_only=True):
        record = row_to_dict(headers, excel_row)
        if any(record.values()):
            rows.append(record)
    return rows


def export_aereo(ws):
    headers = [clean_value(cell.value) for cell in ws[1]]
    headers = headers[: max(index for index, header in enumerate(headers) if header) + 1]
    rows = []
    for excel_row in ws.iter_rows(min_row=3, values_only=True):
        record = row_to_dict(headers, excel_row)
        if any(record.values()):
            rows.append(record)
    return rows


def export_deteccion(ws):
    headers = [clean_value(cell.value) for cell in ws[1]]
    headers = headers[: max(index for index, header in enumerate(headers) if header) + 1]
    resources = []
    people = []
    current_resource = None

    for excel_row in ws.iter_rows(min_row=2, values_only=True):
      record = row_to_dict(headers, excel_row)
      if not any(record.values()):
          continue

      number = record.get("N°", "")
      name = record.get("Torres / Puesto de Observación", "")
      if number:
          current_resource = {
              "N°": number,
              "Región": record.get("Región", ""),
              "Comuna": record.get("Comuna", ""),
              "Recurso": name,
              "Inicio 2° Semestre": record.get("Inicio 2° Semestre", ""),
              "Termino 1° Semestre": record.get("Termino 1° Semestre", ""),
              "Extensión": record.get("Extensión", ""),
              "N° Personas": record.get("N° Personas", ""),
          }
          resources.append(current_resource)
      elif current_resource and name:
          people.append({
              "N° Recurso": current_resource["N°"],
              "Recurso": current_resource["Recurso"],
              "Comuna": current_resource["Comuna"],
              "Nombre Persona": name,
              "Inicio": record.get("Inicio 2° Semestre", current_resource["Inicio 2° Semestre"]),
              "Termino": record.get("Termino 1° Semestre", current_resource["Termino 1° Semestre"]),
          })

    return resources, people


def load_geojson(path):
    with path.open("r", encoding="utf-8-sig") as handle:
        return json.load(handle)


def save_geojson(path, payload):
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        json.dump(payload, handle, ensure_ascii=False, separators=(",", ":"))


def enrich_brigades(terrestre_rows):
    path = DATA_DIR / "brigadas.geojson"
    geojson = load_geojson(path)
    by_unit = {match_key(row.get("Nombre Unidad")): row for row in terrestre_rows if row.get("Nombre Unidad")}
    matched = 0

    for feature in geojson.get("features", []):
        props = feature.setdefault("properties", {})
        key = match_key(props.get("nombre") or props.get("Nombre Unidad"))
        row = by_unit.get(key)
        if not row:
            continue

        matched += 1
        props["poa_region"] = row.get("Región", "")
        props["poa_comuna"] = row.get("Comuna", "")
        props["poa_plan"] = row.get("Plan", "")
        props["poa_proveedor"] = row.get("Proveedor", "")
        props["poa_patente"] = row.get("Patente", "")
        props["personal"] = int(row["Dotación"]) if row.get("Dotación", "").isdigit() else props.get("personal", 0)
        props["base"] = row.get("Comuna") or props.get("base", "")
        props["tipo"] = row.get("Tipo Unidad") or props.get("tipo", "")
        props["clase"] = row.get("Clase") or props.get("clase", "")
        props["tipo_movil"] = row.get("Tipo Movil") or props.get("tipo_movil", "")

    save_geojson(path, geojson)
    return matched


def enrich_towers(deteccion_rows, deteccion_people):
    path = DATA_DIR / "torres.geojson"
    geojson = load_geojson(path)
    by_resource = {match_key(row.get("Recurso")): row for row in deteccion_rows if row.get("Recurso")}
    people_by_resource = {}
    for person in deteccion_people:
        people_by_resource.setdefault(match_key(person.get("Recurso")), []).append(person.get("Nombre Persona", ""))

    matched = 0
    for feature in geojson.get("features", []):
        props = feature.setdefault("properties", {})
        key = match_key(props.get("nombre") or props.get("NOMBRE"))
        row = by_resource.get(key)
        if not row:
            continue

        matched += 1
        props["poa_region"] = row.get("Región", "")
        props["poa_comuna"] = row.get("Comuna", "")
        props["poa_inicio"] = row.get("Inicio 2° Semestre", "")
        props["poa_termino"] = row.get("Termino 1° Semestre", "")
        props["poa_extension"] = row.get("Extensión", "")
        props["poa_personas"] = row.get("N° Personas", "")
        props["poa_nomina"] = "; ".join(name for name in people_by_resource.get(key, []) if name)

    save_geojson(path, geojson)
    return matched


def main():
    wb = load_workbook(POA_PATH, data_only=True, read_only=True)
    terrestre = export_terrestre(wb["POA Terrestre"])
    deteccion, deteccion_personal = export_deteccion(wb["POA Detección "])
    aereo = export_aereo(wb["POA Aéreo"])

    write_csv(METADATA_DIR / "poa_terrestre_2024_2025.csv", terrestre)
    write_csv(METADATA_DIR / "poa_deteccion_recursos_2024_2025.csv", deteccion)
    write_csv(METADATA_DIR / "poa_deteccion_personal_2024_2025.csv", deteccion_personal)
    write_csv(METADATA_DIR / "poa_aereo_2024_2025.csv", aereo)

    summary = {
        "poa_excel": str(POA_PATH.relative_to(ROOT)),
        "csv": {
            "poa_terrestre_2024_2025.csv": len(terrestre),
            "poa_deteccion_recursos_2024_2025.csv": len(deteccion),
            "poa_deteccion_personal_2024_2025.csv": len(deteccion_personal),
            "poa_aereo_2024_2025.csv": len(aereo),
        },
        "matches": {
            "brigadas_geojson": enrich_brigades(terrestre),
            "torres_geojson": enrich_towers(deteccion, deteccion_personal),
        },
    }
    (METADATA_DIR / "poa_procesamiento_resumen.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
